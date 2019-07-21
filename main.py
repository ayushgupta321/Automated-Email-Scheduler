'''

This script requires 4 libraries to work. 
For any missing library or other errors,see README

'''

import datetime
import sys
import time
import smtplib
import openpyxl


smtpObj = smtplib.SMTP('smtp.gmail.com', 587)           #Used SMTP to login to Gmail
smtpObj.ehlo()											# To verify connection
smtpObj.starttls()										# Starting TLS connection
password= input('Enter Password')
smtpObj.login('xyz05556@gmail.com', password)			
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

f= open("data.txt",'r')
date1 = tuple(f.readline())
n=0
i=1
while(1):
	if( sheet.cell(row=i, column=1).value != None ):
	   n=n+1
	   i=i+1
	else:
	   break 


mrng=datetime.datetime(date1)
while(1):
	curtim=datetime.datetime.now()
	if(curtim.hour!=mrng.hour ):
		time.sleep(10000)
	else:
		for i in range(1, n+1, 1):
			if( sheet.cell(row=i, column=2).value == curtim.strftime('%d\%m')):
				name1 = sheet.cell(row=i, column=1).value
				for j in range(1,n+1,1):
					if(i!=j):
						email = sheet.cell(row=j, column=3).value
						name2 = sheet.cell(row=j, column=1).value
						name1 = sheet.cell(row=i, column=1).value
						smtpObj.sendmail('xyz05556@gmail.com', email ,'Subject:' + name1+ '\'s Birthday Today!\nDear '+ name2 +', \n Wish your colleague a very happy birthday.')
				
	break 

smtpObj.quit()
wb.close()
f.close()
